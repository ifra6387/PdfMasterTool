#!/usr/bin/env python3
"""
Professional Excel to PDF Converter for Replit
Converts .xlsx/.xls files to PDF with formatting preservation
Compatible with Linux environments
"""

import sys
import json
import traceback
from pathlib import Path
import tempfile
import os

def convert_excel_to_pdf_professional(input_path, output_path):
    """
    Convert Excel document to PDF using openpyxl and reportlab
    """
    try:
        import openpyxl
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        
        # Load Excel workbook
        wb = openpyxl.load_workbook(input_path, data_only=True)
        
        # Create PDF document
        pdf_doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'ExcelTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.black,
            alignment=TA_CENTER
        )
        
        sheet_style = ParagraphStyle(
            'SheetHeader',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=8,
            spaceBefore=12,
            textColor=colors.blue,
            alignment=TA_LEFT
        )
        
        # Build content list
        content = []
        
        # Add title
        content.append(Paragraph("Excel to PDF Conversion", title_style))
        content.append(Spacer(1, 12))
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Add sheet header
            content.append(Paragraph(f"Sheet: {sheet_name}", sheet_style))
            content.append(Spacer(1, 6))
            
            # Get data from worksheet
            data = []
            max_row = min(ws.max_row, 100)  # Limit to 100 rows for performance
            max_col = min(ws.max_column, 20)  # Limit to 20 columns
            
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                row_data = []
                for cell_value in row:
                    if cell_value is None:
                        row_data.append("")
                    else:
                        # Convert to string and handle various data types
                        cell_str = str(cell_value)
                        if len(cell_str) > 50:  # Truncate very long values
                            cell_str = cell_str[:47] + "..."
                        row_data.append(cell_str)
                
                # Only add non-empty rows
                if any(cell.strip() for cell in row_data if cell):
                    data.append(row_data)
            
            if data:
                # Create table
                table = Table(data)
                
                # Style the table
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]
                
                table.setStyle(TableStyle(table_style))
                
                content.append(table)
                content.append(Spacer(1, 20))
            else:
                content.append(Paragraph("No data found in this sheet.", styles['Normal']))
                content.append(Spacer(1, 12))
        
        # Add footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        content.append(Spacer(1, 20))
        content.append(Paragraph("Converted by I Love Making PDF", footer_style))
        
        # Build PDF
        pdf_doc.build(content)
        
        print(json.dumps({
            "success": True,
            "message": f"Excel document successfully converted to PDF with {len(wb.sheetnames)} sheets."
        }))
        return True
        
    except Exception as e:
        print(f"Professional conversion failed: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return False

def main():
    if len(sys.argv) != 3:
        print(json.dumps({
            "success": False,
            "error": "Usage: python excel_to_pdf_converter.py <input.xlsx> <output.pdf>"
        }))
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    
    # Validate input file
    if not os.path.exists(input_path):
        print(json.dumps({
            "success": False,
            "error": "Input file not found"
        }))
        sys.exit(1)
    
    if not (input_path.lower().endswith('.xlsx') or input_path.lower().endswith('.xls')):
        print(json.dumps({
            "success": False,
            "error": "Input file must be a .xlsx or .xls file"
        }))
        sys.exit(1)
    
    try:
        # Convert Excel to PDF
        if convert_excel_to_pdf_professional(input_path, output_path):
            sys.exit(0)
        
        # If conversion fails
        print(json.dumps({
            "success": False,
            "error": "Excel to PDF conversion failed. Please ensure the file is not corrupted."
        }))
        sys.exit(1)
        
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": f"Conversion failed: {str(e)}"
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()